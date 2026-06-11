import argparse
import pandas as pd
import random
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference

# === FILES ===
reference_run_file = "ref.csv"
current_run_file = "current.csv"
output_file_base = "garmin_analysis"


def get_output_filename(base=output_file_base):
    """Build an output Excel filename with a random 3-digit suffix."""
    return f"{base}_{random.randint(0, 999):03d}.xlsx"


# === HELPERS ===
def extract_date(filename):
    """Extract a run date from a filename.

    The filename is expected to contain a 6-digit date in the form YYMMDD.
    If found, this returns a normalized date string in ISO format YYYY-MM-DD.
    Otherwise, it returns "Unknown".

    Args:
        filename (str): The input filename to parse.

    Returns:
        str: The extracted date string or "Unknown".
    """
    match = re.search(r"(\d{6})", filename)
    if match:
        d = match.group(1)
        return f"20{d[:2]}-{d[2:4]}-{d[4:]}"
    return "Unknown"


def load_file(path):
    """Load a Garmin CSV file and clean lap rows.

    This function reads the CSV into a DataFrame, drops the summary row,
    removes laps shorter than 0.5 kilometers, and resets the index.

    Args:
        path (str): Path to the CSV file.

    Returns:
        pandas.DataFrame: Cleaned lap-level data.
    """
    df = pd.read_csv(path)

    # Remove summary row
    df = df[df["Laps"] != "Summary"]

    # Remove tiny laps (<0.5 km)
    df = df[df["Distance km"] >= 0.5]

    return df.reset_index(drop=True)


def pace_to_seconds(p):
    """Convert a pace string to total seconds.

    The expected input format is "M:SS" where M is minutes and SS is seconds.

    Args:
        p (str): Pace string in minutes and seconds.

    Returns:
        float: Total pace duration in seconds.
    """
    m, s = p.split(":")
    return int(m) * 60 + float(s)


def seconds_to_pace(sec):
    """Convert a seconds value to a pace string.

    If the input is missing or NaN, this returns an empty string.
    Otherwise it formats the value as "M:SS".

    Args:
        sec (float | int): Time in seconds.

    Returns:
        str: Formatted pace string or empty string for missing values.
    """
    if pd.isna(sec):
        return ""
    m = int(sec // 60)
    s = int(round(sec % 60))
    return f"{m}:{s:02d}"


def compare_garmin_data(df1, df2, reference_run_file=reference_run_file, current_run_file=current_run_file):
    """Compare two loaded Garmin dataframes and export the analysis to Excel."""
    output_file = get_output_filename(output_file_base)

    date1 = extract_date(reference_run_file)
    date2 = extract_date(current_run_file)

    # Align lap count
    min_len = min(len(df1), len(df2))
    df1 = df1.iloc[:min_len]
    df2 = df2.iloc[:min_len]

    # Convert pace to seconds
    df1["pace_sec"] = df1["Avg Pace min/km"].apply(pace_to_seconds)
    df2["pace_sec"] = df2["Avg Pace min/km"].apply(pace_to_seconds)

    # === METRICS ===
    metrics = {
        "Pace": ("pace_sec", True),
        "HR": ("Avg HR bpm", False),
        "Power": ("Avg Power W", False),
        "Cadence": ("Avg Run Cadence spm", False),
        "GCT": ("Avg Ground Contact Time ms", True),
        "Stride": ("Avg Stride Length m", False),
        "Vert Osc": ("Avg Vertical Oscillation cm", True),
        "Vert Ratio": ("Avg Vertical Ratio %", True),
    }

    # === BUILD COMPARISON ===
    result = pd.DataFrame()
    result["Lap"] = df1["Laps"]

    for label, (col, lower_is_better) in metrics.items():
        r1 = df1[col].astype(float)
        r2 = df2[col].astype(float)

        diff = r2 - r1
        pct = diff / r1.replace(0, pd.NA)

        if label == "Pace":
            result[f"{label} {date1}"] = r1.apply(seconds_to_pace)
            result[f"{label} {date2}"] = r2.apply(seconds_to_pace)
            result[f"{label} Diff"] = diff.apply(seconds_to_pace)
            result[f"{label} {date1} (sec)"] = r1
            result[f"{label} {date2} (sec)"] = r2
        else:
            result[f"{label} {date1}"] = r1
            result[f"{label} {date2}"] = r2
            result[f"{label} Diff"] = diff

        score = -pct if lower_is_better else pct
        result[f"{label} %"] = pct
        result[f"{label} Score"] = score

    analysis = pd.DataFrame(
        {
            "Lap": df1["Laps"],
            "Pace1_sec": df1["pace_sec"],
            "HR1": df1["Avg HR bpm"],
            "Pace2_sec": df2["pace_sec"],
            "HR2": df2["Avg HR bpm"],
        }
    )

    analysis["Efficiency1"] = analysis["HR1"] / analysis["Pace1_sec"]
    analysis["Efficiency2"] = analysis["HR2"] / analysis["Pace2_sec"]

    summary = pd.DataFrame(
        {
            "Metric": list(metrics.keys()),
            "Avg % Change": [result[f"{m} %"].mean() for m in metrics],
            "Avg Score": [result[f"{m} Score"].mean() for m in metrics],
        }
    )

    summary.loc[len(summary)] = ["Overall Score", "", summary["Avg Score"].mean()]

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Lap Comparison", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        analysis.to_excel(writer, sheet_name="Analysis", index=False)

    wb = load_workbook(output_file)
    ws = wb["Lap Comparison"]

    green = PatternFill(start_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", fill_type="solid")

    for col in ws.iter_cols(min_row=2):
        header = ws.cell(row=1, column=col[0].column).value

        for cell in col:
            if isinstance(cell.value, (int, float)):
                if "Pace" in str(header) and "%" not in str(header):
                    continue

                if "%" in str(header):
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "0.00"

                if "%" in str(header):
                    cell.fill = green if cell.value > 0 else red

    ws2 = wb["Summary"]

    for col in ws2.iter_cols(min_row=2):
        header = ws2.cell(row=1, column=col[0].column).value

        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%" if "%" in str(header) else "0.00"

    headers = [cell.value for cell in ws[1]]
    col1 = headers.index(f"Pace {date1} (sec)") + 1
    col2 = headers.index(f"Pace {date2} (sec)") + 1

    chart = LineChart()
    chart.title = "Pace Comparison (sec)_"

    data = Reference(ws, min_col=col1, max_col=col2, min_row=1, max_row=min_len + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=min_len + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "Z2")

    wb.save(output_file)
    print(f"✅ Analysis complete: {output_file}")


def main(reference_run_file=reference_run_file, current_run_file=current_run_file):
    df1 = load_file(reference_run_file)
    df2 = load_file(current_run_file)
    return compare_garmin_data(df1, df2, reference_run_file, current_run_file)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compare two Garmin run CSV files and export analysis to Excel."
    )
    parser.add_argument(
        "reference_run_file",
        nargs="?",
        default=reference_run_file,
        help="Reference CSV file path (default: ref.csv)",
    )
    parser.add_argument(
        "current_run_file",
        nargs="?",
        default=current_run_file,
        help="Current CSV file path (default: current.csv)",
    )
    args = parser.parse_args()

    main(args.reference_run_file, args.current_run_file)
