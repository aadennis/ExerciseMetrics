import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference

# === FILES ===
file1 = "run_260608_cadence.csv"
file2 = "run_Cadence_260610.csv"
output_file = "garmin_pivot.xlsx"

# === HELPERS ===
def extract_date(filename):
    match = re.search(r'(\d{6})', filename)
    if match:
        d = match.group(1)
        return f"20{d[:2]}-{d[2:4]}-{d[4:]}"
    return "Unknown"


def load_file(path):
    df = pd.read_csv(path)
    df = df[df["Laps"] != "Summary"]
    df = df[df["Distance km"] >= 0.5]
    return df.reset_index(drop=True)


def pace_to_seconds(p):
    m, s = p.split(":")
    return int(m) * 60 + float(s)


def seconds_to_pace(sec):
    if pd.isna(sec):
        return ""
    m = int(sec // 60)
    s = int(round(sec % 60))
    return f"{m}:{s:02d}"


# === LOAD ===
df1 = load_file(file1)
df2 = load_file(file2)

date1 = extract_date(file1)
date2 = extract_date(file2)

min_len = min(len(df1), len(df2))
df1 = df1.iloc[:min_len]
df2 = df2.iloc[:min_len]

df1["pace_sec"] = df1["Avg Pace min/km"].apply(pace_to_seconds)
df2["pace_sec"] = df2["Avg Pace min/km"].apply(pace_to_seconds)

# === METRICS ===
metrics = {
    "Pace": ("pace_sec", True),
    "HR": ("Avg HR bpm", False),
    "Power": ("Avg Power W", False),
    "Cadence": ("Avg Run Cadence spm", False),
    "GCT": ("Avg Ground Contact Time ms", True),
    "Stride": ("Avg Stride Length m", False),
    "Vert Osc": ("Avg Vertical Oscillation cm", True),
    "Vert Ratio": ("Avg Vertical Ratio %", True),
}

# === BUILD COMPARISON ===
result = pd.DataFrame()
result["Lap"] = df1["Laps"]

for label, (col, lower_is_better) in metrics.items():
    r1 = df1[col].astype(float)
    r2 = df2[col].astype(float)

    diff = r2 - r1
    pct = diff / r1.replace(0, pd.NA)

    if label == "Pace":
        # Display columns
        result[f"{label} {date1}"] = r1.apply(seconds_to_pace)
        result[f"{label} {date2}"] = r2.apply(seconds_to_pace)
        result[f"{label} Diff"] = diff.apply(seconds_to_pace)

        # Numeric columns for chart
        result[f"{label} {date1} (sec)"] = r1
        result[f"{label} {date2} (sec)"] = r2
    else:
        result[f"{label} {date1}"] = r1
        result[f"{label} {date2}"] = r2
        result[f"{label} Diff"] = diff

    # Improvement score
    score = -pct if lower_is_better else pct

    result[f"{label} %"] = pct
    result[f"{label} Score"] = score

# === PIVOT (WIDE VIEW) ===
pivot_rows = []
laps = result["Lap"].tolist()

for col in result.columns:
    if col == "Lap":
        continue
    row = [col] + result[col].tolist()
    pivot_rows.append(row)

pivot_df = pd.DataFrame(
    pivot_rows,
    columns=["Metric"] + [f"Lap {lap}" for lap in laps]
)

# === SUMMARY ===
summary = pd.DataFrame({
    "Metric": list(metrics.keys()),
    "Avg % Change": [result[f"{m} %"].mean() for m in metrics],
    "Avg Score": [result[f"{m} Score"].mean() for m in metrics]
})

summary.loc[len(summary)] = [
    "Overall Score",
    "",
    summary["Avg Score"].mean()
]

# === WRITE ===
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    result.to_excel(writer, sheet_name="Lap Comparison", index=False)
    pivot_df.to_excel(writer, sheet_name="Lap Comparison (Wide)", index=False)
    summary.to_excel(writer, sheet_name="Summary", index=False)

# === FORMAT ===
wb = load_workbook(output_file)
ws = wb["Lap Comparison"]

green = PatternFill(start_color="C6EFCE", fill_type="solid")
red = PatternFill(start_color="FFC7CE", fill_type="solid")

for col in ws.iter_cols(min_row=2):
    header = ws.cell(row=1, column=int(col[0].column)).value

    for cell in col:
        if isinstance(cell.value, (int, float)):
            if "%" in str(header):
                cell.number_format = "0.00%"
                cell.fill = green if cell.value > 0 else red
            else:
                cell.number_format = "0.00"

# === CHART (USES NUMERIC PACE) ===
headers = [cell.value for cell in ws[1]]

col1 = headers.index(f"Pace {date1} (sec)") + 1
col2 = headers.index(f"Pace {date2} (sec)") + 1

chart = LineChart()
chart.title = "Pace Comparison (sec)"

data = Reference(ws, min_col=col1, max_col=col2, min_row=1, max_row=min_len + 1)
cats = Reference(ws, min_col=1, min_row=2, max_row=min_len + 1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "Z2")

wb.save(output_file)

print(f"✅ Done: {output_file}")